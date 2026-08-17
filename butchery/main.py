
import os
import secrets
import sqlite3
import random
import string
from datetime import datetime, timedelta
from functools import wraps
import stripe
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_wtf import FlaskForm, CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_talisman import Talisman
from flask_bcrypt import Bcrypt
from wtforms import StringField, PasswordField, IntegerField, FloatField, TextAreaField, SelectField, BooleanField
from wtforms.validators import InputRequired, Email, Length, NumberRange
from flask_wtf.file import FileField, FileAllowed
import openpyxl
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///32slayer_butchery.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['WTF_CSRF_ENABLED'] = True
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

@app.before_request
def check_session_validity():
    if current_user.is_authenticated:
        # Check if user is banned or deleted
        user = User.query.get(current_user.id)
        if not user or user.is_banned:
            logout_user()
            flash('Your session has expired or your account is no longer active.', 'error')
            return redirect(url_for('login'))
        
        # Refresh session to extend lifetime
        session.permanent = True
        app.permanent_session_lifetime = timedelta(minutes=30)

# Initialize extensions
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
csrf = CSRFProtect(app)

# Initialize Limiter
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
)

# Initialize Talisman for security headers
csp = {
    'default-src': [
        '\'self\'',
        'https://cdn.jsdelivr.net',
        'https://fonts.googleapis.com',
        'https://fonts.gstatic.com',
    ],
    'script-src': [
        '\'self\'',
        'https://cdn.jsdelivr.net',
    ],
    'style-src': [
        '\'self\'',
        'https://cdn.jsdelivr.net',
        'https://fonts.googleapis.com',
    ],
}
talisman = Talisman(app, content_security_policy=csp, force_https=False)

# Stripe configuration (Use test keys for now)
# TODO: Replace with your actual Stripe test keys from https://dashboard.stripe.com/test/apikeys
stripe.api_key = os.environ.get('STRIPE_SECRET_KEY', "sk_test_51...")  # Replace with your test secret key
STRIPE_PUBLISHABLE_KEY = os.environ.get('STRIPE_PUBLISHABLE_KEY', "pk_test_51...")  # Replace with your test publishable key

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=True)
    phone = db.Column(db.String(20), unique=True, nullable=True)
    password_hash = db.Column(db.String(120), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    is_banned = db.Column(db.Boolean, default=False)
    failed_login_attempts = db.Column(db.Integer, default=0)
    locked_until = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Product(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    price = db.Column(db.Float, nullable=False)
    stock_quantity = db.Column(db.Integer, nullable=False, default=0)
    category = db.Column(db.String(50))
    image_url = db.Column(db.String(200))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    total_amount = db.Column(db.Float, nullable=False)
    status = db.Column(db.String(20), default='pending')
    stripe_payment_intent_id = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    user = db.relationship('User', backref=db.backref('orders', lazy=True))

class OrderItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id'), nullable=False)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    price = db.Column(db.Float, nullable=False)
    order = db.relationship('Order', backref=db.backref('items', lazy=True))
    product = db.relationship('Product', backref=db.backref('order_items', lazy=True))

class LoginAttempt(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ip_address = db.Column(db.String(45))
    username = db.Column(db.String(80))
    success = db.Column(db.Boolean)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)

class SiteSettings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    background_color = db.Column(db.String(7), default='#ffffff')
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Forms
class LoginForm(FlaskForm):
    username = StringField('Username', validators=[InputRequired(), Length(min=4, max=15)])
    password = PasswordField('Password', validators=[InputRequired(), Length(min=8, max=80)])

class RegistrationForm(FlaskForm):
    username = StringField('Username', validators=[InputRequired(), Length(min=4, max=15)])
    email = StringField('Email', validators=[Email()])
    phone = StringField('Phone Number', validators=[Length(min=10, max=15)])
    password = PasswordField('Password', validators=[InputRequired(), Length(min=8, max=80)])

class ProductForm(FlaskForm):
    name = StringField('Product Name', validators=[InputRequired(), Length(max=100)])
    description = TextAreaField('Description')
    price = FloatField('Price', validators=[InputRequired(), NumberRange(min=0)])
    stock_quantity = IntegerField('Stock Quantity', validators=[InputRequired(), NumberRange(min=0)])
    category = SelectField('Category', choices=[
        ('beef', 'Beef'), ('pork', 'Pork'), ('chicken', 'Chicken'), 
        ('lamb', 'Lamb'), ('processed', 'Processed Meats'), ('other', 'Other')
    ])
    image = FileField('Product Image', validators=[FileAllowed(['jpg', 'jpeg', 'png', 'gif'], 'Images only!')])

class BackgroundForm(FlaskForm):
    background_color = StringField('Background Color', validators=[InputRequired(), Length(min=7, max=7)])

# Login manager
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Utility functions
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Admin access required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def check_account_locked(user):
    """Check if account is locked due to failed login attempts"""
    if user.locked_until and user.locked_until > datetime.utcnow():
        return True
    return False

def log_login_attempt(ip_address, username, success):
    """Log login attempts for security monitoring"""
    attempt = LoginAttempt(
        ip_address=ip_address,
        username=username,
        success=success
    )
    db.session.add(attempt)
    db.session.commit()

def save_uploaded_image(image_file):
    """Save uploaded image and return the filename"""
    if image_file:
        # Create upload directory if it doesn't exist
        upload_dir = app.config['UPLOAD_FOLDER']
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
        
        # Generate unique filename
        filename = secure_filename(image_file.filename)
        name, ext = os.path.splitext(filename)
        unique_filename = f"{name}_{secrets.token_hex(8)}{ext}"
        
        # Save the file
        file_path = os.path.join(upload_dir, unique_filename)
        image_file.save(file_path)
        
        # Return the relative path for storing in database
        return f"uploads/{unique_filename}"
    return None

@app.route('/accept-cookies')
def accept_cookies():
    response = redirect(request.referrer or url_for('index'))
    response.set_cookie('cookies_accepted', 'true', max_age=60*60*24*365, httponly=True, samesite='Lax')
    return response

@app.context_processor
def inject_site_settings():
    """Inject site settings into all templates"""
    settings = SiteSettings.query.first()
    if not settings:
        settings = SiteSettings(background_color='#ffffff')
        db.session.add(settings)
        db.session.commit()
    return dict(site_settings=settings)

# Routes
@app.route('/')
def index():
    products = Product.query.filter_by(is_active=True).all()
    return render_template('index.html', products=products)

@app.route('/catalog')
def catalog():
    category = request.args.get('category')
    # Basic sanitization: only allow alphanumeric categories
    if category and not category.isalnum():
        category = None
    
    if category:
        products = Product.query.filter_by(category=category, is_active=True).all()
    else:
        products = Product.query.filter_by(is_active=True).all()
    
    categories = ['beef', 'pork', 'chicken', 'lamb', 'processed', 'other']
    return render_template('catalog.html', products=products, categories=categories, selected_category=category)

@app.route('/terms-and-conditions')
def terms_conditions():
    return render_template('terms_conditions.html')

@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("3 per hour")
def register():
    form = RegistrationForm()
    if form.validate_on_submit():
        # Check password complexity
        password = form.password.data
        if not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c.isdigit() for c in password):
            flash('Password must contain uppercase, lowercase, and numeric characters.', 'error')
            return render_template('register.html', form=form)
        
        # Check if user exists
        existing_user = User.query.filter(
            (User.username == form.username.data) | 
            (User.email == form.email.data) | 
            (User.phone == form.phone.data)
        ).first()
        
        if existing_user:
            flash('User with this username, email, or phone already exists.', 'error')
            return render_template('register.html', form=form)
        
        # Create new user
        hashed_password = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
        user = User(
            username=form.username.data,
            email=form.email.data if form.email.data else None,
            phone=form.phone.data if form.phone.data else None,
            password_hash=hashed_password
        )
        
        db.session.add(user)
        db.session.commit()
        flash('Registration successful!', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html', form=form)

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        
        if user and check_account_locked(user):
            flash('Account temporarily locked due to failed login attempts.', 'error')
            log_login_attempt(request.remote_addr, form.username.data, False)
            return render_template('login.html', form=form)
        
        if user and user.is_banned:
            flash('Your account has been banned.', 'error')
            log_login_attempt(request.remote_addr, form.username.data, False)
            return render_template('login.html', form=form)
        
        if user and bcrypt.check_password_hash(user.password_hash, form.password.data):
            user.failed_login_attempts = 0
            user.locked_until = None
            db.session.commit()
            login_user(user)
            log_login_attempt(request.remote_addr, form.username.data, True)
            next_page = request.args.get('next')
            # Sanitize next_page to prevent open redirect
            if next_page and not next_page.startswith('/'):
                next_page = url_for('index')
            return redirect(next_page) if next_page else redirect(url_for('index'))
        else:
            if user:
                user.failed_login_attempts += 1
                if user.failed_login_attempts >= 5:
                    user.locked_until = datetime.utcnow() + timedelta(minutes=30)
                db.session.commit()
            
            flash('Invalid username or password.', 'error')
            log_login_attempt(request.remote_addr, form.username.data, False)
    
    return render_template('login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/cart')
def cart():
    cart_items = session.get('cart', {})
    cart_data = []
    total = 0
    
    for product_id, quantity in cart_items.items():
        product = Product.query.get(int(product_id))
        if product:
            subtotal = product.price * quantity
            cart_data.append({
                'product': product,
                'quantity': quantity,
                'subtotal': subtotal
            })
            total += subtotal
    
    return render_template('cart.html', cart_data=cart_data, total=total)

@app.route('/add_to_cart/<int:product_id>')
@login_required
def add_to_cart(product_id):
    # Prevent admin users from adding items to cart
    if current_user.is_admin:
        flash('Admin users cannot add items to cart. Use the admin panel to manage products.', 'warning')
        return redirect(url_for('catalog'))
    
    product = Product.query.get_or_404(product_id)
    cart = session.get('cart', {})
    
    if str(product_id) in cart:
        cart[str(product_id)] += 1
    else:
        cart[str(product_id)] = 1
    
    # Check stock
    if cart[str(product_id)] > product.stock_quantity:
        cart[str(product_id)] = product.stock_quantity
        flash(f'Only {product.stock_quantity} items available in stock.', 'warning')
    
    session['cart'] = cart
    flash(f'{product.name} added to cart!', 'success')
    return redirect(url_for('catalog'))

@app.route('/remove_from_cart/<int:product_id>')
def remove_from_cart(product_id):
    cart = session.get('cart', {})
    if str(product_id) in cart:
        del cart[str(product_id)]
        session['cart'] = cart
        flash('Item removed from cart.', 'info')
    return redirect(url_for('cart'))

@app.route('/checkout')
@login_required
def checkout():
    cart_items = session.get('cart', {})
    if not cart_items:
        flash('Your cart is empty.', 'error')
        return redirect(url_for('cart'))
    
    cart_data = []
    total = 0
    
    for product_id, quantity in cart_items.items():
        product = Product.query.get(int(product_id))
        if product:
            if quantity > product.stock_quantity:
                flash(f'Insufficient stock for {product.name}. Available: {product.stock_quantity}', 'error')
                return redirect(url_for('cart'))
            
            subtotal = product.price * quantity
            cart_data.append({
                'product': product,
                'quantity': quantity,
                'subtotal': subtotal
            })
            total += subtotal
    
    return render_template('checkout.html', cart_data=cart_data, total=total, stripe_key=STRIPE_PUBLISHABLE_KEY)

@app.route('/create_payment_intent', methods=['POST'])
@login_required
def create_payment_intent():
    try:
        cart_items = session.get('cart', {})
        total = 0
        
        for product_id, quantity in cart_items.items():
            product = Product.query.get(int(product_id))
            if product:
                total += product.price * quantity
        
        # Create payment intent
        intent = stripe.PaymentIntent.create(
            amount=int(total * 100),  # Stripe expects cents
            currency='zar',  # South African Rand
            metadata={'user_id': current_user.id}
        )
        
        return jsonify({
            'client_secret': intent.client_secret
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/payment_success')
@login_required
def payment_success():
    payment_intent_id = request.args.get('payment_intent')
    
    if payment_intent_id:
        # Create order
        cart_items = session.get('cart', {})
        total = 0
        
        # Calculate total and create order
        for product_id, quantity in cart_items.items():
            product = Product.query.get(int(product_id))
            if product:
                total += product.price * quantity
        
        order = Order(
            user_id=current_user.id,
            total_amount=total,
            status='completed',
            stripe_payment_intent_id=payment_intent_id
        )
        db.session.add(order)
        db.session.flush()
        
        # Add order items and update stock
        for product_id, quantity in cart_items.items():
            product = Product.query.get(int(product_id))
            if product:
                order_item = OrderItem(
                    order_id=order.id,
                    product_id=product.id,
                    quantity=quantity,
                    price=product.price
                )
                db.session.add(order_item)
                
                # Update stock
                product.stock_quantity -= quantity
        
        db.session.commit()
        
        # Clear cart
        session['cart'] = {}
        
        flash('Payment successful! Your order has been placed.', 'success')
        return render_template('payment_success.html', order=order)
    
    flash('Payment verification failed.', 'error')
    return redirect(url_for('cart'))

# Admin routes
@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    total_users = User.query.count()
    total_products = Product.query.count()
    total_orders = Order.query.count()
    recent_orders = Order.query.order_by(Order.created_at.desc()).limit(10).all()
    
    return render_template('admin/dashboard.html', 
                         total_users=total_users,
                         total_products=total_products,
                         total_orders=total_orders,
                         recent_orders=recent_orders)

@app.route('/admin/products')
@login_required
@admin_required
def admin_products():
    products = Product.query.all()
    return render_template('admin/products.html', products=products)

@app.route('/admin/products/add', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_add_product():
    form = ProductForm()
    if form.validate_on_submit():
        # Handle image upload
        image_url = None
        if form.image.data:
            image_url = save_uploaded_image(form.image.data)
        
        product = Product(
            name=form.name.data,
            description=form.description.data,
            price=form.price.data,
            stock_quantity=form.stock_quantity.data,
            category=form.category.data,
            image_url=image_url
        )
        db.session.add(product)
        db.session.commit()
        flash('Product added successfully!', 'success')
        return redirect(url_for('admin_products'))
    
    return render_template('admin/add_product.html', form=form)

@app.route('/admin/products/edit/<int:product_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_edit_product(product_id):
    product = Product.query.get_or_404(product_id)
    form = ProductForm(obj=product)
    
    if form.validate_on_submit():
        # Handle image upload
        if form.image.data:
            new_image_url = save_uploaded_image(form.image.data)
            if new_image_url:
                # Delete old image if it exists
                if product.image_url:
                    old_image_path = os.path.join('static', product.image_url)
                    if os.path.exists(old_image_path):
                        os.remove(old_image_path)
                product.image_url = new_image_url
        
        product.name = form.name.data
        product.description = form.description.data
        product.price = form.price.data
        product.stock_quantity = form.stock_quantity.data
        product.category = form.category.data
        db.session.commit()
        flash('Product updated successfully!', 'success')
        return redirect(url_for('admin_products'))
    
    return render_template('admin/edit_product.html', form=form, product=product)

@app.route('/admin/products/delete/<int:product_id>')
@login_required
@admin_required
def admin_delete_product(product_id):
    product = Product.query.get_or_404(product_id)
    product.is_active = False
    db.session.commit()
    flash('Product deactivated successfully!', 'success')
    return redirect(url_for('admin_products'))

@app.route('/admin/products/activate/<int:product_id>')
@login_required
@admin_required
def admin_activate_product(product_id):
    product = Product.query.get_or_404(product_id)
    product.is_active = True
    db.session.commit()
    flash('Product activated successfully!', 'success')
    return redirect(url_for('admin_products'))

@app.route('/admin/products/bulk_deactivate')
@login_required
@admin_required
def admin_bulk_deactivate():
    product_ids = request.args.get('ids', '').split(',')
    # Sanitize product_ids: ensure they are all digits
    product_ids = [pid for pid in product_ids if pid.isdigit()]
    
    if product_ids:
        products = Product.query.filter(Product.id.in_(product_ids)).all()
        for product in products:
            product.is_active = False
        db.session.commit()
        flash(f'{len(products)} product(s) deactivated successfully!', 'success')
    return redirect(url_for('admin_products'))

@app.route('/admin/products/bulk_activate')
@login_required
@admin_required
def admin_bulk_activate():
    product_ids = request.args.get('ids', '').split(',')
    # Sanitize product_ids: ensure they are all digits
    product_ids = [pid for pid in product_ids if pid.isdigit()]
    
    if product_ids:
        products = Product.query.filter(Product.id.in_(product_ids)).all()
        for product in products:
            product.is_active = True
        db.session.commit()
        flash(f'{len(products)} product(s) activated successfully!', 'success')
    return redirect(url_for('admin_products'))

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    users = User.query.all()
    return render_template('admin/users.html', users=users)

@app.route('/admin/users/ban/<int:user_id>')
@login_required
@admin_required
def admin_ban_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_admin:
        flash('Cannot ban admin users.', 'error')
    else:
        user.is_banned = True
        db.session.commit()
        flash(f'User {user.username} has been banned.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/unban/<int:user_id>')
@login_required
@admin_required
def admin_unban_user(user_id):
    user = User.query.get_or_404(user_id)
    user.is_banned = False
    db.session.commit()
    flash(f'User {user.username} has been unbanned.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/export_data')
@login_required
@admin_required
def export_data():
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Users sheet
    ws_users = wb.active
    ws_users.title = "Users"
    ws_users.append(["ID", "Username", "Email", "Phone", "Is Admin", "Is Banned", "Created At"])
    
    users = User.query.all()
    for user in users:
        ws_users.append([
            user.id, user.username, user.email, user.phone,
            user.is_admin, user.is_banned, user.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    # Products sheet
    ws_products = wb.create_sheet("Products")
    ws_products.append(["ID", "Name", "Description", "Price", "Stock", "Category", "Is Active"])
    
    products = Product.query.all()
    for product in products:
        ws_products.append([
            product.id, product.name, product.description, product.price,
            product.stock_quantity, product.category, product.is_active
        ])
    
    # Orders sheet
    ws_orders = wb.create_sheet("Orders")
    ws_orders.append(["ID", "User ID", "Username", "Total Amount", "Status", "Created At"])
    
    orders = Order.query.join(User).all()
    for order in orders:
        ws_orders.append([
            order.id, order.user_id, order.user.username,
            order.total_amount, order.status, order.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    # Save file
    filename = f"highveld_data_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    return send_file(filename, as_attachment=True)

@app.route('/admin/export_purchases')
@login_required
@admin_required
def export_purchases():
    # Create Excel workbook for purchases
    wb = openpyxl.Workbook()
    
    # Purchase Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Purchase Summary"
    ws_summary.append([
        "Order ID", "Customer", "Customer Email", "Customer Phone",
        "Order Date", "Total Amount", "Status", "Payment Intent ID", "Item Count"
    ])
    
    orders = Order.query.join(User).all()
    for order in orders:
        item_count = len(order.items)
        ws_summary.append([
            order.id,
            order.user.username,
            order.user.email or "N/A",
            order.user.phone or "N/A",
            order.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            order.total_amount,
            order.status,
            order.stripe_payment_intent_id or "N/A",
            item_count
        ])
    
    # Detailed Purchase Items sheet
    ws_details = wb.create_sheet("Purchase Details")
    ws_details.append([
        "Order ID", "Customer", "Product Name", "Product Category",
        "Quantity", "Unit Price", "Total Price", "Order Date", "Order Status"
    ])
    
    order_items = OrderItem.query.join(Order).join(User).join(Product).all()
    for item in order_items:
        ws_details.append([
            item.order_id,
            item.order.user.username,
            item.product.name,
            item.product.category,
            item.quantity,
            item.price,
            item.quantity * item.price,
            item.order.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            item.order.status
        ])
    
    # Sales Analytics sheet
    ws_analytics = wb.create_sheet("Sales Analytics")
    ws_analytics.append([
        "Product Name", "Category", "Total Quantity Sold", "Total Revenue", "Average Order Size"
    ])
    
    # Calculate sales analytics
    products = Product.query.all()
    for product in products:
        total_quantity = sum(item.quantity for item in product.order_items)
        total_revenue = sum(item.quantity * item.price for item in product.order_items)
        order_count = len(set(item.order_id for item in product.order_items))
        avg_order_size = total_quantity / order_count if order_count > 0 else 0
        
        ws_analytics.append([
            product.name,
            product.category,
            total_quantity,
            total_revenue,
            round(avg_order_size, 2)
        ])
    
    # Monthly Sales sheet
    ws_monthly = wb.create_sheet("Monthly Sales")
    ws_monthly.append(["Month", "Total Orders", "Total Revenue", "Average Order Value"])
    
    # Group orders by month
    monthly_data = {}
    for order in orders:
        month_key = order.created_at.strftime("%Y-%m")
        if month_key not in monthly_data:
            monthly_data[month_key] = {"orders": 0, "revenue": 0}
        monthly_data[month_key]["orders"] += 1
        monthly_data[month_key]["revenue"] += order.total_amount
    
    for month, data in sorted(monthly_data.items()):
        avg_order_value = data["revenue"] / data["orders"] if data["orders"] > 0 else 0
        ws_monthly.append([
            month,
            data["orders"],
            data["revenue"],
            round(avg_order_value, 2)
        ])
    
    # Save file
    filename = f"highveld_purchases_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    
    return send_file(filename, as_attachment=True)

@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_settings():
    form = BackgroundForm()
    settings = SiteSettings.query.first()
    
    if not settings:
        settings = SiteSettings(background_color='#ffffff')
        db.session.add(settings)
        db.session.commit()
    
    if form.validate_on_submit():
        color = form.background_color.data
        # Validate hex color format
        if color.startswith('#') and len(color) == 7:
            try:
                int(color[1:], 16)  # Validate hex format
                settings.background_color = color
                settings.updated_at = datetime.utcnow()
                db.session.commit()
                flash('Background color updated successfully!', 'success')
                return redirect(url_for('admin_settings'))
            except ValueError:
                flash('Invalid color format. Please use hex format (e.g., #ffffff)', 'error')
        else:
            flash('Invalid color format. Please use hex format (e.g., #ffffff)', 'error')
    
    form.background_color.data = settings.background_color
    return render_template('admin/settings.html', form=form, settings=settings)

@app.route('/security/logins')
@login_required
@admin_required
def security_logins():
    attempts = LoginAttempt.query.order_by(LoginAttempt.timestamp.desc()).limit(100).all()
    return render_template('admin/security_logins.html', attempts=attempts)

# Error handlers
@app.errorhandler(404)
def not_found(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('errors/500.html'), 500

@app.errorhandler(403)
def forbidden(error):
    return render_template('errors/403.html'), 403

# Initialize database and create admin user
def init_db():
    with app.app_context():
        db.create_all()
        
        # Create admin user if it doesn't exist
        admin = User.query.filter_by(username='Siphelele23').first()
        if not admin:
            admin_password = bcrypt.generate_password_hash('Admin123!').decode('utf-8')
            admin = User(
                username='Siphelele23',
                email='admin@highveld.co.za',
                phone='0825883367',
                password_hash=admin_password,
                is_admin=True
            )
            db.session.add(admin)
        
        # Add sample products if none exist
        if Product.query.count() == 0:
            sample_products = [
                Product(name='Premium Beef Steak', description='High-quality beef steak, perfect for grilling', price=89.99, stock_quantity=50, category='beef'),
                Product(name='Free Range Chicken Breast', description='Fresh, free-range chicken breast fillets', price=45.99, stock_quantity=30, category='chicken'),
                Product(name='Lamb Chops', description='Tender lamb chops, ideal for braai', price=129.99, stock_quantity=25, category='lamb'),
                Product(name='Pork Ribs', description='Succulent pork ribs with perfect marbling', price=69.99, stock_quantity=40, category='pork'),
                Product(name='Boerewors', description='Traditional South African sausage', price=39.99, stock_quantity=60, category='processed'),
            ]
            
            for product in sample_products:
                db.session.add(product)
        
        # Initialize site settings if not exists
        if SiteSettings.query.count() == 0:
            default_settings = SiteSettings(background_color='#ffffff')
            db.session.add(default_settings)
        
        db.session.commit()
        print("Database initialized successfully!")

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=False)
